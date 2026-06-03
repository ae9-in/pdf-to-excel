import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from models import ConversionResult

# ── Colour palette ────────────────────
C = {
    "dark":    "1E1E2E", "accent":  "4F46E5", "accent2": "7C3AED",
    "light":   "F5F3FF", "mid":     "E0E7FF", "text":    "374151",
    "muted":   "6B7280", "white":   "FFFFFF", "red":     "EF4444",
    "yellow":  "F59E0B", "green":   "10B981", "border":  "C7D2FE",
}

TEAM_COLORS = {
    "HR": "818CF8", 
    "NEW JOINERS": "A78BFA",
    "BUSINESS OPERATIONS": "34D399", 
    "PRODUCT": "FB923C",
    "SPOS": "F472B6", 
    "SOLID": "60A5FA",
    "STUDENT RELATED BUS": "FBBF24", 
    "SPECIFIC": "6EE7B7",
}

def _fill(h): 
    return PatternFill("solid", fgColor=h)

def _font(h=None, bold=False, sz=9, italic=False):
    return Font(name="Calibri", color=h or C["text"], bold=bold, size=sz, italic=italic)

def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _brd():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_merged_border(ws, cell_range):
    """Ensure borders are applied to all cells within a merged range."""
    s = Side(style="thin", color=C["border"])
    border = Border(left=s, right=s, top=s, bottom=s)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def generate_excel(result: ConversionResult) -> bytes:
    """
    Generate a fully-formatted Excel workbook from a ConversionResult.
    Returns bytes.
    """
    wb = Workbook()
    
    # Active sheet is sheet 1
    _build_task_sheet(wb, result)
    _build_dashboard_sheet(wb, result)
    _build_field_guide_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_task_sheet(wb: Workbook, result: ConversionResult):
    ws = wb.active
    ws.title = "Task Allocation"
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A":5, "B":26, "C":22, "D":28, "E":11, "F":9, "G":14, "H":52, "I":28, "J":14, "K":13, "L":13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── HEADER ROWS ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    _apply_merged_border(ws, "A1:K1")
    ws["A1"] = f"TASK ALLOCATION REPORT  —  {result.organisation.upper()}"
    ws["A1"].font = _font(C["white"], True, 17)
    ws["A1"].fill = _fill(C["dark"])
    ws["A1"].alignment = _aln("center", "center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:K2")
    _apply_merged_border(ws, "A2:K2")
    ws["A2"] = f"{result.sprint_title}  |  Converted via EarlyBird PDF-to-Excel Engine"
    ws["A2"].font = _font("A5B4FC", False, 9, False)
    ws["A2"].fill = _fill(C["dark"])
    ws["A2"].alignment = _aln("center", "center")
    ws.row_dimensions[2].height = 18

    # Statistics row styling
    last_row = 4 + len(result.tasks)
    ws.row_dimensions[3].height = 20
    stats = [
        ("A3", "Total Tasks:"), ("B3", f'=COUNTA(B5:B{last_row})'),
        ("D3", "High Priority:"), ("E3", f'=COUNTIF(E5:E{last_row},"High")'),
        ("G3", "On Leave:"), ("H3", f'=COUNTIF(J5:J{last_row},"Leave")'),
        ("J3", "Completed:"), ("K3", f'=COUNTIF(J5:J{last_row},"Completed")'),
    ]
    # Fill row 3 cells with background color
    for col_idx in range(1, 12):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = _fill("2D2B55")

    for addr, val in stats:
        c = ws[addr]
        c.value = val
        is_label = addr[0] in "ADGJ"
        c.font = _font("A5B4FC" if is_label else C["white"], True, 9)
        c.alignment = _aln("right" if is_label else "center", "center")

    # Column headers
    headers = ["#", "Name", "Team", "Task Title", "Priority", "SP", "Due Date", "Description", "Email", "Status", "Done %"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = _font(C["white"], True, 10)
        c.fill = _fill(C["accent"])
        c.alignment = _aln("center", "center")
        c.border = _brd()
    ws.row_dimensions[4].height = 22

    # Data validation dropdowns
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked,Leave"', allow_blank=True)
    dv_status.sqref = "J5:J300"
    ws.add_data_validation(dv_status)

    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv_priority.sqref = "E5:E300"
    ws.add_data_validation(dv_priority)

    # ── DATA ROWS ───────────────────────────────────────────────────────────
    pri_color = {"High": C["red"], "Medium": C["yellow"], "Low": C["green"]}

    for i, task in enumerate(result.tasks, start=5):
        ws.row_dimensions[i].height = 26
        bg = "FAFAFA" if i % 2 == 0 else C["white"]
        row_vals = [
            task.id, task.name, task.team, task.task_title,
            task.priority, task.story_points, task.due_date,
            task.description, task.email, task.status,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill = _fill(bg)
            c.border = _brd()
            # Alignments: Center for Col 1, 6, 7. Left for others. wrap text for description.
            c.alignment = _aln("center" if ci in (1, 6, 7) else "left", "center", ci == 8)
            c.font = _font(C["text"], False, 9)
            
        # Completion % (Done %)
        pct_cell = ws.cell(row=i, column=11, value=task.completion_pct / 100.0)
        pct_cell.number_format = "0%"
        pct_cell.fill = _fill(bg)
        pct_cell.border = _brd()
        pct_cell.alignment = _aln("center", "center")
        pct_cell.font = _font(C["text"], False, 9)

        # Team colour (Col 3)
        tc = TEAM_COLORS.get(task.team.upper().strip(), "94A3B8")
        ws.cell(row=i, column=3).font = _font(tc, True, 9)

        # Priority colour (Col 5)
        pc = pri_color.get(task.priority, C["text"])
        ws.cell(row=i, column=5).font = _font(pc, True, 9)

    if last_row >= 5:
        # Conditional formatting
        def cf(h): 
            return PatternFill("solid", fgColor=h)

        # Priority Formatting rules
        ws.conditional_formatting.add(f"E5:E{last_row}", CellIsRule("equal", ['"High"'], fill=cf("FEE2E2"), font=_font(C["red"], True)))
        ws.conditional_formatting.add(f"E5:E{last_row}", CellIsRule("equal", ['"Medium"'], fill=cf("FEF3C7"), font=_font(C["yellow"], True)))
        ws.conditional_formatting.add(f"E5:E{last_row}", CellIsRule("equal", ['"Low"'], fill=cf("D1FAE5"), font=_font(C["green"], True)))

        # Status Formatting rules
        ws.conditional_formatting.add(f"J5:J{last_row}", CellIsRule("equal", ['"Completed"'], fill=cf("D1FAE5"), font=_font(C["green"], True)))
        ws.conditional_formatting.add(f"J5:J{last_row}", CellIsRule("equal", ['"Blocked"'], fill=cf("FEE2E2"), font=_font(C["red"], True)))
        ws.conditional_formatting.add(f"J5:J{last_row}", CellIsRule("equal", ['"Leave"'], fill=cf("F3F4F6"), font=Font(color="9CA3AF", italic=True, name="Calibri", size=9)))
        
        ws.auto_filter.ref = f"A4:K{last_row}"


def _build_dashboard_sheet(wb: Workbook, result: ConversionResult):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A":26, "B":13, "C":13, "D":13, "E":13, "F":13, "G":13, "H":13, "I":13, "J":13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    heights = {1: 30, 2: 16, 3: 24, 4: 30, 5: 30, 6: 24, 7: 30, 8: 30, 9: 24, 10: 6, 11: 22, 12: 20}
    for row, h in heights.items():
        ws.row_dimensions[row].height = h

    last_row = 4 + len(result.tasks)

    # ── HEADER ROWS ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    _apply_merged_border(ws, "A1:J1")
    ws["A1"] = f"SPRINT DASHBOARD  —  {result.organisation.upper()}"
    ws["A1"].font = _font(C["white"], True, 15)
    ws["A1"].fill = _fill(C["dark"])
    ws["A1"].alignment = _aln("center", "center")

    ws.merge_cells("A2:J2")
    _apply_merged_border(ws, "A2:J2")
    ws["A2"] = "All formulas auto-update when Task Allocation sheet is edited"
    ws["A2"].font = _font("A5B4FC", False, 9, False)
    ws["A2"].fill = _fill(C["dark"])
    ws["A2"].alignment = _aln("center", "center")

    # ── STATS CARDS ─────────────────────────────────────────────────────────
    if len(result.tasks) == 0:
        cards = [
            ("A4:C4", "A5:C5", "Total Tasks", 0, C["accent"]),
            ("D4:F4", "D5:F5", "High Priority", 0, C["red"]),
            ("G4:I4", "G5:I5", "In Progress", 0, C["yellow"]),
            ("A7:C7", "A8:C8", "Completed", 0, C["green"]),
            ("D7:F7", "D8:F8", "On Leave", 0, C["muted"]),
            ("G7:I7", "G8:I8", "Unique Teams", 0, C["accent2"]),
        ]
    else:
        cards = [
            # (Label range, Value range, Label text, Formula, background hex)
            ("A4:C4", "A5:C5", "Total Tasks", f"=COUNTA('Task Allocation'!B5:B{last_row})", C["accent"]),
            ("D4:F4", "D5:F5", "High Priority", f"=COUNTIF('Task Allocation'!E5:E{last_row},\"High\")", C["red"]),
            ("G4:I4", "G5:I5", "In Progress", f"=COUNTIF('Task Allocation'!J5:J{last_row},\"In Progress\")", C["yellow"]),
            ("A7:C7", "A8:C8", "Completed", f"=COUNTIF('Task Allocation'!J5:J{last_row},\"Completed\")", C["green"]),
            ("D7:F7", "D8:F8", "On Leave", f"=COUNTIF('Task Allocation'!J5:J{last_row},\"Leave\")", C["muted"]),
            ("G7:I7", "G8:I8", "Unique Teams", f"=SUMPRODUCT(1/COUNTIF('Task Allocation'!C5:C{last_row},'Task Allocation'!C5:C{last_row}))", C["accent2"]),
        ]

    for lbl_range, val_range, lbl_text, formula, bg in cards:
        # Fill colors (before merging to avoid MergedCell exceptions)
        for row in ws[lbl_range]:
            for cell in row:
                cell.fill = _fill(bg)
        for row in ws[val_range]:
            for cell in row:
                cell.fill = _fill(bg)

        ws.merge_cells(lbl_range)
        _apply_merged_border(ws, lbl_range)
        ws.merge_cells(val_range)
        _apply_merged_border(ws, val_range)
        
        lbl_cell = ws[lbl_range.split(":")[0]]
        lbl_cell.value = lbl_text
        lbl_cell.font = _font(C["white"], False, 9)
        lbl_cell.alignment = _aln("center", "top")
        
        val_cell = ws[val_range.split(":")[0]]
        val_cell.value = formula
        val_cell.font = _font(C["white"], True, 24)
        val_cell.alignment = _aln("center", "center")

    # ── TEAM BREAKDOWN TABLE ────────────────────────────────────────────────
    ws["A11"] = "Team Breakdown"
    ws["A11"].font = _font(C["dark"], True, 12)

    headers = ["Team", "Tasks", "High", "In Progress", "Completed", "On Leave"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=12, column=ci, value=h)
        c.font = _font(C["white"], True, 10)
        c.fill = _fill(C["accent"])
        c.alignment = _aln("center", "center")
        c.border = _brd()

    # Dynamic extraction of teams from tasks
    unique_teams = sorted(list(set(task.team.strip() for task in result.tasks if task.team and task.team.strip())))
    if not unique_teams:
        unique_teams = ["HR", "BUSINESS OPERATIONS", "PRODUCT", "SPOS"]

    teams_data = []
    for team in unique_teams:
        color = TEAM_COLORS.get(team.upper(), "94A3B8")
        teams_data.append((team, color))

    for idx, (team, color) in enumerate(teams_data, start=13):
        ws.row_dimensions[idx].height = 22
        bg = "FAFAFA" if idx % 2 == 0 else C["white"]
        
        # Col 1: Team Name
        c1 = ws.cell(row=idx, column=1, value=team)
        c1.font = _font(color, True, 9)
        c1.alignment = _aln("left", "center")
        c1.fill = _fill(bg)
        c1.border = _brd()
        
        # Col 2: Tasks Count
        c2 = ws.cell(row=idx, column=2, value=f"=COUNTIF('Task Allocation'!C5:C{last_row}, A{idx})")
        c2.font = _font(C["text"], True, 9)
        c2.alignment = _aln("center", "center")
        c2.fill = _fill(bg)
        c2.border = _brd()
        
        # Col 3: High Priority Count
        c3 = ws.cell(row=idx, column=3, value=f"=COUNTIFS('Task Allocation'!C5:C{last_row}, A{idx}, 'Task Allocation'!E5:E{last_row}, \"High\")")
        c3.font = _font(C["text"], True, 9)
        c3.alignment = _aln("center", "center")
        c3.fill = _fill(bg)
        c3.border = _brd()
        
        # Col 4: In Progress Count
        c4 = ws.cell(row=idx, column=4, value=f"=COUNTIFS('Task Allocation'!C5:C{last_row}, A{idx}, 'Task Allocation'!J5:J{last_row}, \"In Progress\")")
        c4.font = _font(C["text"], True, 9)
        c4.alignment = _aln("center", "center")
        c4.fill = _fill(bg)
        c4.border = _brd()
        
        # Col 5: Completed Count
        c5 = ws.cell(row=idx, column=5, value=f"=COUNTIFS('Task Allocation'!C5:C{last_row}, A{idx}, 'Task Allocation'!J5:J{last_row}, \"Completed\")")
        c5.font = _font(C["text"], True, 9)
        c5.alignment = _aln("center", "center")
        c5.fill = _fill(bg)
        c5.border = _brd()
        
        # Col 6: On Leave Count
        c6 = ws.cell(row=idx, column=6, value=f"=COUNTIFS('Task Allocation'!C5:C{last_row}, A{idx}, 'Task Allocation'!J5:J{last_row}, \"Leave\")")
        c6.font = _font(C["text"], True, 9)
        c6.alignment = _aln("center", "center")
        c6.fill = _fill(bg)
        c6.border = _brd()


def _build_field_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("Field Guide")
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A":28, "B":22, "C":13, "D":32, "E":22, "F":30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 22
    for r in range(4, 14):
        ws.row_dimensions[r].height = 20

    # ── HEADER ROWS ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    _apply_merged_border(ws, "A1:F1")
    ws["A1"] = "PDF to EXCEL FIELD GUIDE  |  EarlyBird India Conversion System"
    ws["A1"].font = _font(C["white"], True, 13)
    ws["A1"].fill = _fill(C["dark"])
    ws["A1"].alignment = _aln("center", "center")

    # Table Headers
    headers = ["PDF COLUMN", "EXCEL COLUMN", "DATA TYPE", "VALIDATION RULE", "EXAMPLE VALUE", "NOTES"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = _font(C["white"], True, 10)
        c.fill = _fill(C["accent"])
        c.alignment = _aln("left", "center")
        c.border = _brd()

    # Guide Rows Data
    guide_rows = [
        ("Person name heading", "B — Name", "Text", "Required, non-empty", "Madhumati Angadi", "Full name extracted from PDF"),
        ("Team / Department label", "C — Team", "Text", "Must match team list", "HR", "Standardised team name"),
        ("Task bold heading", "D — Task Title", "Text", "Max 60 chars", "Interview Scheduling", "Short task summary"),
        ("Priority badge / keyword", "E — Priority", "Dropdown", "High / Medium / Low", "High", "Parse from colour or keyword"),
        ("Story Points (if present)", "F — Story Points", "Number", "0-13 integer", "5", "Default 3 if absent"),
        ("Date in PDF header", "G — Due Date", "Date", "DD-MMM-YY", "29-May-26", "Parse from sprint heading"),
        ("Bullet points / body", "H — Description", "Long Text", "Max 500 chars", "Schedule interviews…", "Concat bullets, clean whitespace"),
        ("Email (if shown)", "I — Email", "Email", "Must contain @", "m@gmail.com", "Leave blank if not in PDF"),
        ("AI-inferred or manual", "J — Status", "Dropdown", "Not Started / In Progress / Completed / Blocked / Leave", "In Progress", "Infer from wording: 'leave'→Leave"),
        ("Manual input", "K — Done %", "Percentage", "0-100", "60%", "Default 0; update manually")
    ]

    for idx, row in enumerate(guide_rows, start=4):
        bg = "F5F3FF" if idx % 2 == 0 else C["white"]
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=idx, column=col_idx, value=val)
            c.font = _font(C["text"], False, 9)
            c.fill = _fill(bg)
            c.border = _brd()
            c.alignment = _aln("left", "center", True)
